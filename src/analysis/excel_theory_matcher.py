"""Excel理论匹配引擎 - 专门针对Excel统计文件中的理论进行匹配"""

from typing import List, Dict, Optional, Set
from pathlib import Path
import re
from loguru import logger

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl未安装,Excel理论匹配功能将不可用")


class ExcelTheoryMatcher:
    """Excel理论匹配器 - 从Excel统计文件中提取理论并进行匹配"""

    def __init__(self, excel_dirs: List[str] = None):
        """
        初始化Excel理论匹配器

        Args:
            excel_dirs: Excel文件目录列表,默认为项目根目录下的FDC系列文件夹
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("无法初始化ExcelTheoryMatcher: openpyxl未安装")
            self.theories_set = set()
            self.theory_sources = {}
            return

        self.excel_dirs = excel_dirs or []
        self.theories_set: Set[str] = set()  # 所有唯一理论的集合
        self.theory_sources: Dict[str, List[Dict]] = {}  # 理论来源信息

        # 如果未指定目录,自动扫描FDC系列目录
        if not self.excel_dirs:
            self._auto_discover_excel_dirs()

        # 加载所有Excel中的理论
        self._load_theories_from_excel()

    def _auto_discover_excel_dirs(self):
        """自动发现FDC系列Excel目录"""
        base_path = Path(__file__).parent.parent.parent  # 项目根目录

        # 查找所有FDC-XX文件夹
        fdc_dirs = sorted(base_path.glob("FDC-*"))

        for fdc_dir in fdc_dirs:
            if fdc_dir.is_dir():
                # 查找案例统计子目录
                stats_dirs = list(fdc_dir.glob("*案例统计*"))
                if stats_dirs:
                    for stats_dir in stats_dirs:
                        self.excel_dirs.append(str(stats_dir))
                else:
                    # 如果没有统计子目录,直接使用FDC目录
                    self.excel_dirs.append(str(fdc_dir))

        logger.info(f"发现 {len(self.excel_dirs)} 个Excel目录")

    def _load_theories_from_excel(self):
        """从Excel文件中加载理论"""
        if not OPENPYXL_AVAILABLE:
            return

        total_theories = 0

        for excel_dir in self.excel_dirs:
            excel_path = Path(excel_dir)

            # 查找所有Excel文件
            excel_files = list(excel_path.glob("*.xlsx"))

            for excel_file in excel_files:
                # 跳过临时文件
                if excel_file.name.startswith("~$"):
                    continue

                try:
                    theories = self._extract_theories_from_file(excel_file)
                    total_theories += len(theories)
                    logger.debug(f"  {excel_file.name}: {len(theories)} 个理论")
                except Exception as e:
                    logger.warning(f"  加载 {excel_file.name} 失败: {e}")

        logger.success(f"✓ 从Excel加载了 {len(self.theories_set)} 个唯一理论 (总计 {total_theories} 条记录)")

    def _extract_theories_from_file(self, excel_file: Path) -> List[str]:
        """
        从单个Excel文件中提取理论

        Args:
            excel_file: Excel文件路径

        Returns:
            理论列表
        """
        theories = []

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active

            # 查找理论列 (通常在第18列,但我们也可以通过表头查找)
            theory_col_idx = None

            # 遍历前几行查找包含"理论"的表头
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for idx, cell_value in enumerate(row):
                    if cell_value and isinstance(cell_value, str):
                        if "教学思维理论" in cell_value or "理论" in cell_value:
                            theory_col_idx = idx
                            break
                if theory_col_idx is not None:
                    break

            # 如果没找到,默认使用第18列(索引17)
            if theory_col_idx is None:
                theory_col_idx = 17

            # 提取理论数据
            for row in ws.iter_rows(min_row=3, values_only=True):  # 从第3行开始(跳过表头)
                if len(row) > theory_col_idx:
                    theory_value = row[theory_col_idx]

                    if theory_value and isinstance(theory_value, str):
                        theory_value = theory_value.strip()

                        if theory_value and len(theory_value) > 1:  # 过滤空值和单字符
                            theories.append(theory_value)
                            self.theories_set.add(theory_value)

                            # 记录来源
                            if theory_value not in self.theory_sources:
                                self.theory_sources[theory_value] = []

                            self.theory_sources[theory_value].append({
                                'file': excel_file.name,
                                'case_name': row[1] if len(row) > 1 else 'N/A',  # 案例名称通常在第2列
                                'case_code': row[2] if len(row) > 2 else 'N/A'   # 案例编号通常在第3列
                            })

            wb.close()

        except Exception as e:
            logger.error(f"提取理论失败: {excel_file.name} - {e}")

        return theories

    def match_theories(self, input_theories: List[str]) -> Dict[str, Dict]:
        """
        匹配输入理论与Excel中的理论 (支持标准化匹配)

        Args:
            input_theories: 输入的理论列表 (已标准化)

        Returns:
            匹配结果字典
        """
        if not OPENPYXL_AVAILABLE:
            return {}

        logger.info(f"Excel理论匹配: 检查 {len(input_theories)} 个理论...")

        # 导入标准化工具
        try:
            from ..utils.theory_normalizer import TheoryNormalizer
        except ImportError:
            TheoryNormalizer = None

        matched_results = {}

        for theory in input_theories:
            # 首先尝试精确匹配
            if theory in self.theories_set:
                sources = self.theory_sources.get(theory, [])
                matched_results[theory] = {
                    'input_theory': theory,
                    'matched_theory': theory,
                    'match_type': 'exact',
                    'source_count': len(sources),
                    'sources': sources[:5],
                    'excel_coverage': True
                }
                logger.debug(f"  ✓ Excel精确匹配: {theory} (出现{len(sources)}次)")
                continue

            # 如果精确匹配失败,尝试用标准化方式查找Excel中的变体
            matched_count = 0
            all_sources = []

            if TheoryNormalizer:
                # 获取该标准理论的所有变体
                variants = TheoryNormalizer.get_all_variants(theory)

                # 在Excel中查找所有变体
                for variant in variants:
                    if variant in self.theories_set:
                        sources = self.theory_sources.get(variant, [])
                        matched_count += len(sources)
                        all_sources.extend(sources)

                if matched_count > 0:
                    matched_results[theory] = {
                        'input_theory': theory,
                        'matched_theory': theory,
                        'match_type': 'normalized',  # 新类型:通过标准化匹配
                        'source_count': matched_count,
                        'sources': all_sources[:5],
                        'excel_coverage': True,
                        'matched_variants': [v for v in variants if v in self.theories_set]
                    }
                    logger.debug(f"  ✓ Excel标准化匹配: {theory} (通过变体匹配到{matched_count}次)")
                    continue

            # 如果标准化匹配也失败,尝试传统模糊匹配
            fuzzy_match = self._fuzzy_match_excel_theory(theory)

            if fuzzy_match:
                sources = self.theory_sources.get(fuzzy_match, [])
                matched_results[theory] = {
                    'input_theory': theory,
                    'matched_theory': fuzzy_match,
                    'match_type': 'fuzzy',
                    'source_count': len(sources),
                    'sources': sources[:5],
                    'excel_coverage': True
                }
                logger.debug(f"  ≈ Excel模糊匹配: {theory} -> {fuzzy_match} (出现{len(sources)}次)")
            else:
                # 未在Excel中找到
                matched_results[theory] = {
                    'input_theory': theory,
                    'matched_theory': None,
                    'match_type': 'not_found',
                    'source_count': 0,
                    'sources': [],
                    'excel_coverage': False
                }
                logger.debug(f"  ✗ Excel未找到: {theory}")

        logger.success(f"✓ Excel匹配完成: {sum(1 for r in matched_results.values() if r['excel_coverage'])} 个理论在Excel中")

        return matched_results

    def _fuzzy_match_excel_theory(self, input_theory: str) -> Optional[str]:
        """
        模糊匹配Excel中的理论

        Args:
            input_theory: 输入理论

        Returns:
            匹配到的Excel理论,如果没有则返回None
        """
        # 标准化输入
        normalized_input = re.sub(r'\s+', ' ', input_theory).strip()
        normalized_input = normalized_input.replace('(', '(').replace(')', ')')
        normalized_input = normalized_input.replace('(', '(').replace(')', ')')

        # 策略1: 子串包含匹配
        for excel_theory in self.theories_set:
            normalized_excel = re.sub(r'\s+', ' ', excel_theory).strip()
            normalized_excel = normalized_excel.replace('(', '(').replace(')', ')')
            normalized_excel = normalized_excel.replace('(', '(').replace(')', ')')

            # 双向检查
            if normalized_input.lower() in normalized_excel.lower() or \
               normalized_excel.lower() in normalized_input.lower():
                return excel_theory

        # 策略2: 提取中英文分别匹配
        chinese_part = re.sub(r'[^\u4e00-\u9fff]', '', input_theory).strip()
        english_part = re.sub(r'[^a-zA-Z]', ' ', input_theory).strip()
        english_part = re.sub(r'\s+', ' ', english_part).strip()

        for excel_theory in self.theories_set:
            excel_chinese = re.sub(r'[^\u4e00-\u9fff]', '', excel_theory).strip()
            excel_english = re.sub(r'[^a-zA-Z]', ' ', excel_theory).strip()
            excel_english = re.sub(r'\s+', ' ', excel_english).strip()

            # 中文部分匹配
            if chinese_part and excel_chinese and chinese_part == excel_chinese:
                return excel_theory

            # 英文部分匹配
            if english_part and excel_english:
                if english_part.lower() == excel_english.lower():
                    return excel_theory

        return None

    def get_all_excel_theories(self) -> List[str]:
        """获取所有Excel中的理论列表"""
        return sorted(list(self.theories_set))

    def get_theory_sources(self, theory: str) -> List[Dict]:
        """获取理论的所有来源"""
        return self.theory_sources.get(theory, [])

    def get_statistics(self) -> Dict:
        """获取Excel理论统计信息"""
        return {
            'total_unique_theories': len(self.theories_set),
            'excel_directories': len(self.excel_dirs),
            'most_frequent_theories': self._get_most_frequent_theories(10)
        }

    def _get_most_frequent_theories(self, top_n: int = 10) -> List[Dict]:
        """获取出现最频繁的理论"""
        theory_counts = [
            {
                'theory': theory,
                'count': len(sources)
            }
            for theory, sources in self.theory_sources.items()
        ]

        return sorted(theory_counts, key=lambda x: x['count'], reverse=True)[:top_n]
